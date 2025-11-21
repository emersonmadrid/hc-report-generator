from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook
import unicodedata
from datetime import date, datetime


def _normalizar(texto: str) -> str:
    """
    Convierte a mayúsculas, quita tildes y espacios extra.
    Ej: 'Calificación del registro' -> 'CALIFICACION DEL REGISTRO'
    """
    if texto is None:
        return ""
    # Quitar tildes
    nfkd = unicodedata.normalize("NFD", str(texto))
    sin_tildes = "".join(c for c in nfkd if unicodedata.category(c) != "Mn")
    # Mayúsculas y sin espacios sobrantes
    return " ".join(sin_tildes.upper().split())


def _encontrar_hoja_correcta(wb):
    """
    Busca la hoja que contenga la tabla con los datos evaluados.
    Primero intenta buscar hojas con nombres específicos,
    luego busca por contenido.
    """
    # Prioridad 1: Buscar hojas con nombres específicos
    hojas_prioritarias = ['ejemplo -variasHC', 'evalúa anexo5 -lleno 1HC', 'ejemplo', 'evalua']
    
    for nombre_hoja in hojas_prioritarias:
        for sheet_name in wb.sheetnames:
            if nombre_hoja.lower() in sheet_name.lower():
                print(f"  → Usando hoja por nombre: {sheet_name}")
                return wb[sheet_name]
    
    # Prioridad 2: Buscar por contenido específico de la tabla de evaluación
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Buscar si contiene los títulos de la tabla de evaluación
        for row in range(1, min(50, ws.max_row + 1)):
            for col in range(1, min(10, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    texto_normalizado = _normalizar(cell_value)
                    # Buscar "N° de HC evaluada" o similar
                    if ("N" in texto_normalizado and "HC" in texto_normalizado and "EVALUADA" in texto_normalizado):
                        print(f"  → Usando hoja por contenido: {sheet_name}")
                        return ws
    
    # Prioridad 3: Buscar el formato general
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        for row in range(1, min(20, ws.max_row + 1)):
            for col in range(1, min(10, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    texto_normalizado = _normalizar(cell_value)
                    if "FORMATO" in texto_normalizado and "EVALUACION" in texto_normalizado and "CALIDAD" in texto_normalizado:
                        print(f"  → Usando hoja por formato: {sheet_name}")
                        return ws
    
    return None


def parse_auditoria_excel(path: Path) -> Dict[str, List[str]]:
    """
    Lee un Excel con el formato de auditoría y devuelve los datos
    necesarios para llenar una tabla del informe.

    Devuelve un diccionario con:
      - hc: lista de códigos de historia clínica
      - fechas: lista de fechas (string)
      - porcentajes: lista de % cumplimiento
      - calificaciones: lista de calificación del registro
    """
    wb = load_workbook(path, data_only=True)
    
    print(f"\n=== Procesando archivo: {path.name} ===")
    print(f"Hojas disponibles: {wb.sheetnames}")
    
    # Buscar la hoja correcta
    ws = _encontrar_hoja_correcta(wb)
    
    if ws is None:
        raise ValueError(
            f"No se encontró la hoja con el formato esperado en {path.name}. "
            f"Se buscó una hoja que contenga 'FORMATO DE EVALUACIÓN DE LA CALIDAD DE REGISTO EN CONSULTA EXTERNA'"
        )
    
    print(f"✓ Hoja encontrada: {ws.title}")

    fila_hc = fila_fecha = fila_porc = fila_calif = None
    col_inicio = None

    print("\nBuscando las filas clave en TODO el Excel...")
    # Buscar en todas las celdas las filas que contienen estos textos
    for row in range(1, min(200, ws.max_row + 1)):
        for col in range(1, min(20, ws.max_column + 1)):
            raw_value = ws.cell(row=row, column=col).value
            txt = _normalizar(raw_value)

            if not txt:
                continue

            # Mostrar solo las primeras filas para no saturar el log
            if row <= 20 and col <= 3:
                print(f"  Fila {row}, Col {col}: {raw_value}")

            # Buscar los textos específicos del Excel
            # HC: "CODIFICACIÓN DE LA HISTORIA CLÍNICA"
            if "CODIFICACION" in txt and "HISTORIA" in txt and "CLINICA" in txt:
                fila_hc = row
                if col_inicio is None:
                    col_inicio = col
                print(f"  ✓✓✓ Fila HC encontrada en: fila={row}, col={col} - '{raw_value}'")
            
            # Fecha: "FECHA DE LA ATENCIÓN BRINDADA"
            if "FECHA" in txt and "ATENCION" in txt and "BRINDADA" in txt:
                fila_fecha = row
                if col_inicio is None:
                    col_inicio = col
                print(f"  ✓✓✓ Fila FECHA encontrada en: fila={row}, col={col} - '{raw_value}'")
            
            # Porcentaje: "PORCENTAJE DE CUMPLIMIENTO ALCANZADO"
            if "PORCENTAJE" in txt and "CUMPLIMIENTO" in txt and "ALCANZADO" in txt:
                fila_porc = row
                if col_inicio is None:
                    col_inicio = col
                print(f"  ✓✓✓ Fila PORCENTAJE encontrada en: fila={row}, col={col} - '{raw_value}'")

    # La calificación la calcularemos basándonos en el porcentaje
    # >= 90%: SATISFACTORIO
    # 75-89%: POR MEJORAR
    # < 75%: DEFICIENTE
    print(f"  → Calificación se calculará automáticamente según el porcentaje")

    if None in (fila_hc, fila_fecha, fila_porc):
        raise ValueError(
            f"No se encontraron todas las filas requeridas en {path.name}. "
            f"Encontradas - HC: {'✓' if fila_hc else '✗'}, "
            f"Fecha: {'✓' if fila_fecha else '✗'}, "
            f"Porcentaje: {'✓' if fila_porc else '✗'}"
        )

    print(f"\n✓ Todas las filas encontradas:")
    print(f"  HC en fila: {fila_hc}")
    print(f"  Fecha en fila: {fila_fecha}")
    print(f"  Porcentaje en fila: {fila_porc}")
    print(f"  Columna de inicio (títulos): {col_inicio}")

    # Encontrar la primera columna con datos (después de la columna de títulos)
    # Buscamos en la fila de HC la primera columna con contenido
    primera_col_con_datos = None
    print(f"\nBuscando primera columna con datos después de col {col_inicio}...")
    for c in range(col_inicio + 1, ws.max_column + 1):
        val = ws.cell(row=fila_hc, column=c).value
        if val is not None and str(val).strip() != "":
            primera_col_con_datos = c
            print(f"  ✓ Primera columna con datos encontrada: columna {c}")
            break
    
    if primera_col_con_datos is None:
        raise ValueError(f"No se encontraron datos en las filas después de la columna {col_inicio}")
    
    col_datos = primera_col_con_datos

    # Tomamos los datos desde la columna siguiente a los títulos
    col = col_datos
    hc_list: List[str] = []
    fecha_list: List[str] = []
    porc_list: List[str] = []
    calif_list: List[str] = []

    print(f"\nExtrayendo datos desde la columna {col_datos}...")
    print(f"  Explorando hasta la columna {ws.max_column}...")
    
    # Primero ver qué hay en todas las columnas de estas filas
    print(f"\n  Contenido de la fila HC (fila {fila_hc}):")
    for c in range(1, min(20, ws.max_column + 1)):
        val = ws.cell(row=fila_hc, column=c).value
        if val:
            print(f"    Col {c}: {val}")
    
    print(f"\n  Contenido de la fila FECHA (fila {fila_fecha}):")
    for c in range(1, min(20, ws.max_column + 1)):
        val = ws.cell(row=fila_fecha, column=c).value
        if val:
            print(f"    Col {c}: {val}")
    
    print(f"\n  Contenido de la fila PORCENTAJE (fila {fila_porc}):")
    for c in range(1, min(20, ws.max_column + 1)):
        val = ws.cell(row=fila_porc, column=c).value
        if val:
            print(f"    Col {c}: {val}")
    
    print(f"\nExtrayendo datos desde columna {col_datos}:")
    while col <= ws.max_column:
        hc_val = ws.cell(row=fila_hc, column=col).value
        fecha_val = ws.cell(row=fila_fecha, column=col).value
        porc_val = ws.cell(row=fila_porc, column=col).value

        print(f"  Columna {col}: HC={hc_val}, Fecha={fecha_val}, %={porc_val}")

        # Si ya no hay HC, asumimos que se acabaron las columnas útiles
        if hc_val in (None, ""):
            break

        hc_list.append(str(hc_val).strip())

        # Fecha → string amigable
        if isinstance(fecha_val, (date, datetime)):
            fecha_list.append(fecha_val.strftime("%d-%b-%y"))
        elif fecha_val is None:
            fecha_list.append("")
        else:
            fecha_list.append(str(fecha_val).strip())

        porc_list.append("" if porc_val is None else str(porc_val).strip())
        
        # Calcular calificación según porcentaje
        calif_val = ""
        if porc_val is not None:
            try:
                # Extraer el número del porcentaje (puede venir como "85%" o "85" o 0.85)
                porc_str = str(porc_val).replace("%", "").strip()
                porc_num = float(porc_str)
                
                # Si el porcentaje está en formato decimal (0.85), convertir a porcentaje
                if porc_num < 1:
                    porc_num = porc_num * 100
                
                # Asignar calificación
                if porc_num >= 90:
                    calif_val = "SATISFACTORIO"
                elif porc_num >= 75:
                    calif_val = "POR MEJORAR"
                else:
                    calif_val = "DEFICIENTE"
            except (ValueError, AttributeError):
                calif_val = ""
        
        calif_list.append(calif_val)

        col += 1

    print(f"\n✓ Datos extraídos:")
    print(f"  - {len(hc_list)} historias clínicas: {hc_list}")
    print(f"  - Fechas: {fecha_list}")
    print(f"  - Porcentajes: {porc_list}")
    print(f"  - Calificaciones: {calif_list}")

    return {
        "hc": hc_list,
        "fechas": fecha_list,
        "porcentajes": porc_list,
        "calificaciones": calif_list,
    }